import zipfile
import os
import re
import hashlib
import requests
import openpyxl
# Задание номер 1

directory_to_extract_to = 'D:\\python'
arch_file = 'D:\\python\\tiff-4.2.0_lab1.zip'
zip_file = zipfile.ZipFile(arch_file)
zip_file.extractall(directory_to_extract_to)

# Задание номер 2

name_file = zip_file.namelist()
txt_file = []
for r, d, f in os.walk(directory_to_extract_to):
    for name in f:
        if name.endswith(".sh"):
            txt_file.append(os.path.join(r, name))
for file in txt_file:
    fop = open(file, "rb").read()
    print(file + '   ' + str(hashlib.md5(fop).hexdigest()))

# Задание номер 3

target_hash = "4636f9ae9fef12ebd56cd39586d33cfb"
target_file = ''
target_file_data = ''
tmp = 1
for r, d, f in os.walk(directory_to_extract_to):
    for name in f:
        target_file = os.path.join(r,name)
        target_file_data = open(target_file,"rb").read()
        if str(hashlib.md5(target_file_data).hexdigest()) == target_hash:
            print(target_file)
            print(target_file_data)
            tmp = 0
            break
    if tmp == 0:
        break

# Задание номер 4

r = requests.get(target_file_data)
result_dct = {}
counter = 0
lines = re.findall(r'<div class="Table-module_row__3TH83">.*?</div>.*?</div>.*?</div>.*?</div>.*?</div>', r.text)
Table = list()
for line in lines:
    # извлечение заголовков таблицы
    if counter == 0:
        # Удаление тегов
        headers = re.sub('<.*?>', ';', line)
        headers = re.sub(r'\(\+[0-9\s]+\)', '', headers)
        while headers.find(';;') != -1:
            headers = re.sub(';;', ';', headers)
        headers = headers[1:len(headers)-1]
        Table.append(headers)

N = int(input("Input number of row: "))
temp = Table[N]

tmp_split = temp.split(";")
country_name = tmp_split[0]
col1_val = re.sub("\xa0", '', tmp_split[1])
col2_val = re.sub("\xa0", '', tmp_split[2])
col3_val = re.sub(r"(\xa0|\*)", '', tmp_split[3])
if tmp_split[4] == '_':
    col4_val = -1
else:
    col4_val = re.sub("\xa0", '', tmp_split[4])

result_dct = dict()
result_dct["Country"] = country_name[4:]
result_dct["Sick"] = int(col1_val)
result_dct["Died"] = int(col2_val)
result_dct["Recovered"] = int(col3_val)
result_dct["Active case"] = int(col4_val)
print(result_dct)

#задание №5
row = len(Table)
column = 5
file_name = input("Input file name to save: ")
wb = openpyxl.Workbook()
ws = wb.active
first = Table[0]
Name = first.split(';')
for i in range(1, column):
    a = Name[i-1]
    ws.cell(column=i+1,row=1, value=a)
last = Table[row-1]
last_split = last.split(';')
for i in range(1, row-1):
    tmp = Table[i]
    temp = tmp.split(';')
    for j in range(0, column):
        v = temp[j]
        ws.cell(column=j+1, row=i+1, value=v)
    wb.save(file_name)
for j in range(0, column):
    v = last_split[j+1]
    ws.cell(column=j+1,row=row,value=v)
    wb.save(file_name)


#задание №6
target_country = input("Введите название страны: ")
for i in range(1, len(Table)):
    tmp = Table[i]
    tmp_split = tmp.split(';')
    country = tmp_split[0]
    if(target_country == country[4:]):
        col1 = re.sub("\xa0", '', tmp_split[1])
        col2 = re.sub("\xa0", '', tmp_split[2])
        col3 = re.sub(r"(\xa0|\*)", '', tmp_split[3])
        if tmp_split[4] == '_':
            col4 = -1
        else:
            col4 = re.sub("\xa0", '', tmp_split[4])
        result = dict()
        result["Sick"] = int(col1)
        result["Died"] = int(col2)
        result["Recovered"] = int(col3)
        result["Active case"] = int(col4)
        print(result)